import pandas as pd
import openpyxl

xlspath =  "株レーダー.xlsx"
workbook = openpyxl.load_workbook(xlspath)
stname = "list"
sheet = workbook[stname]

def open_kabubook():
    workbook = openpyxl.load_workbook(xlspath)

####################################################
#   listシートから銘柄コードのリストを取得する
####################################################
def get_kabulist(kabulist):
    # 値の入っている最大行、最大列まで列単位でセルの値を取得する
    for cols in sheet.iter_cols(min_row=3, min_col=2, max_col=2):
        for cell in cols:
            if cell.value not in kabulist:
                kabulist.append(cell.value)

def del_sheet(codes):
    for cd in codes:
        sht_code = str(cd)
        if sht_code in workbook.sheetnames :
            # 既存シートを削除
            del workbook[sht_code]

    # Excelファイルを保存
    workbook.save(xlspath)
    workbook.close()


def write_kabudata(df, code):
    sht_code = str(code)
    check = False
    print(df.head(1))


    if sht_code in workbook.sheetnames :
        # 既存シートを削除
        del workbook[sht_code]
        # Excelファイルを保存
        workbook.save(xlspath)

    with pd.ExcelWriter(xlspath, mode="a") as writer:
        df.to_excel(writer, sheet_name=sht_code)


def save_wkbook():
    workbook.save(xlspath)

def close_wkbook():
    # ロードしたExcelファイルを閉じる
    workbook.close()
