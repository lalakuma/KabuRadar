# 解析後に出力したCSVファイルを集計する
from numpy import empty
import pandas as pd
import os
import glob
import openpyxl as px
import copy
import getConfig
import sqlight as db
from openpyxl.styles import Font, PatternFill
from openpyxl.formatting.rule import CellIsRule, FormulaRule
import getConfig as conf
import win32com.client as win32
import os
win32c = win32.constants
#################################################################
# トレード銘柄決定処理
#################################################################
def decide_trade(DirPath):
    # DBに接続
    conn, cursor = db.connect_db()
    df_set = db.read_rec_all(conn, cursor, "tbl_code_set")
    # DBクローズ
    db.close_db(conn)

    # フォルダ内の全ファイルを取得
    allFiles = glob.glob(DirPath + "\*.csv") # 指定したフォルダーの全エクセルファイルを変数に代入します
    df = pd.DataFrame()
    list_ = []

    # 結果ファイルがなければ終了
    if len(allFiles) == 0:
        return df, list_

    # フォルダ内のファイルを全て処理する
    for file_ in allFiles:
        df = pd.read_csv(file_, encoding="ms932", sep=",")    
        # 文字列「code」を検索
        poscode = file_.find("code")
        # 「code」がないファイルはパス
        if poscode == -1:
            continue
        code = file_[poscode+4:poscode+8]
        df["code"] = code
        list_.append(df)
    
    if len(list_) == 0:
        print("集計対象ファイルが見つかりません")
        return df, list_
    # 結合
    df_con = pd.concat(list_, join='inner') # joinをinnerに指定
    # NaNのある行を削除
    df_con = df_con.dropna(axis = 0, how = 'any')
    df_con = df_con.sort_values(['Index', 'code'])

    # 並べ替え
    df_con = df_con.reset_index()
    df_con = pd.merge(df_con, df_set, on='code')            # tbl_code_setとマージ
    df_con = df_con.sort_values("Index", ascending=True)    # 日付順にソート
    df_con = df_con.loc[:,["Index","code","open","close","PF","mark","buygain","sellgain","income"]]
    df_con = df_con.reset_index(drop=True)
    #選別　（お試し用）
    daytotal = 0
    wktotal = 0
    lst_data = []

    if len(df_con) == 0:
        print("No data")
        return

    date = df_con.iat[0,0]
    samelen = 0
    lst_tdyBuy = []     # 当日購入リスト
    lst_PreBuy = []     # 前営業日購入リスト
    for row in range(len(df_con)):
        # 同一日付で処理済の行は読み飛ばす
        if samelen > 0:
            samelen -= 1
            continue
        # 日付が一致するものを抽出
        df_samedate = (df_con[df_con["Index"].str.contains(date[:10])])
        df_samedate = df_samedate.sort_values("PF", ascending=False)    # 高PF順

        # 同一日付内でループ
        for samerow in df_samedate.itertuples():
            # 条件によってmarkを変更したいのでタプルではなく辞書に変換
            dic_row = samerow._asdict()
            #--------------------#
            # エントリー側の判定
            #--------------------#
            if dic_row["mark"] == "新買" or dic_row["mark"] == "新売":
                # 現在値を加算していく
                wktotal += dic_row["close"]                # 一時トータルに終値加算

                # トータル10000(買値が100万円)以内なら購入対象として追加
                if wktotal < 10000:                     # 一時トータルで比較
#                if wktotal < 5000:                     # 一時トータルで比較
                    lst_tdyBuy.append(dic_row["code"])  # 当日購入リストにコードを追加
                    lst_data.append(dic_row)            # データリストに追加
                    daytotal += dic_row["close"]        # 正式トータルに追加                
                wktotal = daytotal  # 一時トータル更新

            #--------------------#
            # 決済側の判定
            #--------------------#
            # 前日リストの中にcodeがあるか検索（前日購入済みの株かどうか）
            findcode = dic_row["code"] in lst_PreBuy
            # 見つけたら集計リストに追加
            if findcode == True:
                # 当日購入リストにあるか判定
                tdyfindcode = dic_row["code"] in lst_tdyBuy
                # 既に購入リスト側に追加されている場合は重複になってしまうので追加しない
                if tdyfindcode == False:
                    lst_data.append(dic_row)
                    dic_row['mark'] = "決済"       # 購入はないので決済のみに書き換え
                else:
                    dic_row['mark'] = "決済+" + dic_row['mark']
        
            #----------------------------#
            # 決済がなければ利益を0にする
            #----------------------------#
            if dic_row["mark"] == "新買" or dic_row["mark"] == "新売":
                dic_row["buygain"] = 0
                dic_row["sellgain"] = 0

        #----------------------------------
        # 次の日付グループに向けての準備
        #----------------------------------
        # 当日購入リストを前日購入リストにコピー
        lst_PreBuy = copy.copy(lst_tdyBuy)        
        lst_tdyBuy.clear()
        daytotal = 0
        wktotal = 0
        samelen = len(df_samedate) - 1
        nextpos = row + samelen + 1
        # 終了判定
        if nextpos < len(df_con):
            # 次の日付をセット
            date = df_con.iat[nextpos, 0]
        else:
            break
    return df_con, lst_data
    
#################################################################
# 集計ファイル作成
#################################################################
def shuukei_makeExl(shuukei_path, stance):
    fileShukei = "集計_" + stance + ".xlsx"
    # トレード決定処理
    df_con, lst_data = decide_trade(shuukei_path)
    if len(lst_data) == 0:
        print("集計対象ファイルがありません。")
        return 0,"",0

    filepath = shuukei_path + fileShukei
    # エクセルに書き込み
    df_con.to_excel(filepath, sheet_name='全コード結合', encoding="shift_jis")

    dfreal = pd.DataFrame(lst_data)    
    with pd.ExcelWriter(filepath, engine="openpyxl", mode="a") as writer:
        dfreal.to_excel(writer, sheet_name='1日10000以下', index=False)

    # エクセルファイルの先頭行にフィルターをつける
    if(len(df_con) > 0):
        wb= px.load_workbook(filepath)          # [Sheet1]
        ws = wb.active
        ws.column_dimensions['B'].width =22
        ws.column_dimensions['F'].width =10
        ws.auto_filter.ref = ws.dimensions
        ws['B1'] = 'date'
        for row, cellObj in enumerate(list(ws.columns)[9]): # セル(J列)に累積計算関数を書き込む
            if row == 0:
                continue
            n = '=IF(ISNUMBER(J' + str(row) + '),J' + str(row) + '+H' + str(row+1) + '+I' + str(row+1) + ',0)'
            cellObj.value = n

    if(len(dfreal) > 0):
        ws1 = wb.worksheets[1]                  # [Sheet2]
        ws1.column_dimensions['B'].width =22
        ws1.column_dimensions['F'].width =10
        ws1.auto_filter.ref = ws1.dimensions
        ws1['B1'] = 'date'
        for row, cellObj in enumerate(list(ws1.columns)[9]): # セル(J列)に累積計算関数を書き込む
            if row == 0:
                continue
            n = '=IF(ISNUMBER(J' + str(row) + '),J' + str(row) + '+H' + str(row+1) + '+I' + str(row+1) + ',0)'
            cellObj.value = n

            ws1.cell(row+1,column=11).value = '=YEAR(B' + str(row+1) + ')'
            ws1.cell(row+1,column=12).value = '=MONTH(B' + str(row+1) + ')'

        ws1.cell(row=1,column=11).value = '年'
        ws1.cell(row=1,column=12).value = '月'

        # セルの色を設定
        fmtarea = 'F2:F' + str(row + 1)
        gray_fill = PatternFill(bgColor='FF99FF', fill_type='solid')
        formula_rule1 = FormulaRule(formula=['$F2="新買"'], fill=gray_fill)
        formula_rule2 = FormulaRule(formula=['$F2="決済+新買"'], fill=gray_fill)
        ws1.conditional_formatting.add(fmtarea, formula_rule1)
        ws1.conditional_formatting.add(fmtarea, formula_rule2)

    if(len(df_con) > 0):
        wb.save(filepath)
    
    # 最終利益を集計ファイルより取得
    
    finalRieki=0

    return lst_data, filepath, finalRieki


def shuukei_toCsv(shuukei_path):
    filelst = os.listdir(path=shuukei_path)
    df = pd.DataFrame()

    wk_pf = 0.0
    #fl_pf = 0.0
    wk_pgain = 0
    total_pgain = 0
    total_mgain = 0
    wk_mgain = 0
    filecnt = 0
    cnt_win = 0
    cnt_lose = 0
    strpath = ""

    for fl in filelst:
        if fl[:4] == "code":
            arr = fl.split('_')
            wk_pf = float((arr[5])[2:]) 

            # 利益と損失を取得
            wk_pgain = int((arr[6])[2:])
            total_pgain += wk_pgain
            wk_mgain = abs(int(((arr[7])[:-4])[2:]))
            total_mgain += wk_mgain

            # win lose の累積件数を取得
            strwl = arr[2]
            lpos = strwl.find('L')
            cnt_win += int(strwl[1:lpos])
            cnt_lose += int(strwl[lpos + 1:])
            dic = {'code':arr[1], 
                            'code':(arr[0])[4:], 
                            'rsi':(arr[1])[3:], 
                            'winlose':strwl, 
                            'winPer':int((arr[3])[:-1]), 
                            'incomes':int((arr[4])[3:]), 
                            'pg':wk_pgain, 
                            'mg':wk_mgain, 
                            'pf':wk_pf}
            dfdic = pd.DataFrame(dic, index=['i',])
            df = pd.concat( [df, dfdic], ignore_index=True)

    if len(df) > 0 :
        print(df)
        #日付をインデックスにして、必要なアイテム順に並び替え
        df = df.set_index("code").loc[:,["pf","pg","mg","incomes","winlose","winPer","rsi"]]
        fWinPer = (cnt_win / (cnt_win + cnt_lose)) * 100
#        strWinPer = "rate" + '{:.1f}'.format(df['winPer'].mean())
        strWinPer = "rate" + '{:.2f}'.format(fWinPer)
        print(strWinPer)
        strincome = "all" + '{:}'.format(int(df['incomes'].sum()))
        print(strincome)
        if total_mgain != 0:
            strpfall = "PF" + str(round(total_pgain / total_mgain, 3))
        else:
            strpfall = "PF" + str(round(total_pgain))

        strpath = shuukei_path + strpfall + "_" + "W" + str(cnt_win) + "L" + str(cnt_lose) + "_" + strWinPer + "_" + strincome + ".csv"
        df.to_csv(strpath, encoding="shift_jis")

    return strpath    

#################################################################
# ピボットテーブル作成
# 引数：skfilepath　集計ファイルのパス
# 戻り値：なし
#################################################################
def create_pivottable(skfilepath):
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    excel.Visible = True  

    ## 読み込み時に絶対パスを指定しなければエラーになる
    fpath = os.path.join(os.getcwd(),skfilepath)
    wb = excel.Workbooks.Open(fpath)

    ## Sheet 1 指定し､フィルターを有効にする
#    wbs1 = wb.Sheets('-下')
    wbs1 = wb.Sheets('1日10000以下')

    ## ピボットテーブルの作成
    wbs2_name = 'pivot'
    wb.Sheets.Add().Name = wbs2_name
    wbs2 = wb.Sheets(wbs2_name)
    pvt_name = 'pvt'
    pc = wb.PivotCaches().Create(SourceType=win32c.xlDatabase, SourceData=wbs1.UsedRange)
    pc.CreatePivotTable(TableDestination='{sheet}!R3C1'.format(sheet=wbs2_name), TableName=pvt_name)

    ## ピボットテーブルの設定
    wbs2.PivotTables(pvt_name).PivotFields('年').Orientation = win32c.xlRowField
    wbs2.PivotTables(pvt_name).PivotFields('月').Orientation = win32c.xlColumnField
    wbs2.PivotTables(pvt_name).PivotFields('buygain').Orientation = win32c.xlDataField

    wbs2.Cells(4, 1).Select()

    ## ファイルを閉じる
    wb.Close(True)
    excel.Quit()

#shuukei_toCsv('..\\..\\output\\honban\\')
# sklst, skfilepath, finalRieki = shuukei_makeExl('..\\..\\output\\honban\\', 'TST')
# create_pivottable(skfilepath)





